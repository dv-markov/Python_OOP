# Загрузка структуры объекта из файла excel

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from dataclasses import dataclass

wb = load_workbook("config.xlsx", read_only=True)
print(wb.sheetnames)
# ws = wb.active
# ws = wb['Structure']
ws = wb[wb.sheetnames[0]]
print(ws)

# решение через объекты openpyxl
wsl1 = [row for row in ws]
print(wsl1)
print(len(wsl1))
for row in wsl1:
    print(row[0].value, row[1].value)
row_indx = 0
while wsl1[row_indx][0].value != '#':
    row_indx += 1
print(row_indx)

# решение через фактические значения ячеек
wsl2 = tuple(tuple(cell.value for cell in row) for row in ws)
wb.close()
print(wsl2)
row_indx = 0
while wsl2[row_indx][0] != '#':
    row_indx += 1
print(row_indx)

# в итоговом кортеже остается только структура данных
wsl = wsl2[row_indx:]
print(wsl)


# @dataclass()
class ValvePart:
    def __init__(self, level, name, parent=None):
        self.level = level
        self.name = name
        self.parent = parent
        self.bom = dict()
        self.params = dict()

    def add_item(self, name, obj=None):
        self.bom[name] = obj

    def add_self(self):
        self.parent.add_item(self.name, self)

    def get_parent(self, depth=1):
        if depth == 1:
            return self.parent
        elif depth > 1:
            return self.parent.get_parent(depth - 1)

    def set_params(self, parameter_list):
        p = parameter_list.split(', ')
        self.params = dict.fromkeys(p)

    def __repr__(self):
        return f"\n{(self.level - 1) * '-'} L:{self.level} - {self.name}" \
               f"{('; Параметры: ' + str(tuple(self.params.keys())) + '; Структура BOM: ') if self.params else ' '}" \
               f"{list(x for x in self.bom.values()) if self.bom else ''}"


valve_list = []
vp_last = None
for row in wsl:
    r = row[0].strip() if row[0] else ''
    if '#' in r:
        level = r.count('#')
        vp = ValvePart(level, row[1].strip())

        if level == 1:
            params = row[2].strip()
            vp.set_params(params)
            valve_list.append(vp)
            # vp_last = vp
        elif level > vp_last.level:
            vp.parent = vp_last
            vp.add_self()
            # vp_last = vp
        elif level == vp_last.level:
            vp.parent = vp_last.get_parent()
            vp.add_self()
            # vp_last = vp
        elif level < vp_last.level:
            vp.parent = vp_last.get_parent(vp_last.level - (level - 1))
            vp.add_self()
        vp_last = vp

print(valve_list)

# for valve in valve_list:
#     print(valve.__dict__)

# print(*(f'L:{p.level} - {p.name}' for p in valve_list), sep='; ')






