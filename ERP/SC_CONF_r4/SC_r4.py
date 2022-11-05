# Загрузка структуры объекта и перечня компонентов из файла excel

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from collections.abc import Iterable

wb = load_workbook("config.xlsx", read_only=True, data_only=True)
print(wb.sheetnames)
# ws = wb.active
ws = wb['Templates']
# ws = wb[wb.sheetnames[0]]

# решение через фактические значения ячеек
wsl2 = tuple(tuple(cell.value for cell in row) for row in ws)
# wb.close()
print(wsl2)
row_indx = 0
while wsl2[row_indx][0] != '#':
    row_indx += 1
print(row_indx)

# в итоговом кортеже остается только структура данных
wsl = wsl2[row_indx:]
print(wsl)


# @dataclass()
class ValvePartTemplate:
    def __init__(self, level, name, parent=None):
        self.level = level
        self.name = name
        self.parent = parent
        self.bom = dict()
        self.params = dict()

    def add_item(self, name, obj=None):
        self.bom[name] = obj

    def add_self(self):
        self.parent.add_item(self.name, self)

    def get_parent(self, depth=1):
        if depth == 1:
            return self.parent
        elif depth > 1:
            return self.parent.get_parent(depth - 1)

    def set_params(self, parameter_list):
        p = parameter_list.split(', ')
        self.params = dict.fromkeys(p)

    def get_params(self):
        return self.params

    def __repr__(self):
        return f"\n{(self.level - 1) * '-'} L:{self.level} - {self.name}" \
               f"{('; Параметры: ' + str(tuple(self.params.keys())) + '; Структура BOM: ') if self.params else ' '}" \
               f"{list(x for x in self.bom.values()) if self.bom else ''}"


# создание шаблонов BOM
valve_list = []
vp_last = None
for row in wsl:
    r = row[0].strip() if row[0] else ''
    if '#' in r:
        level = r.count('#')
        vp = ValvePartTemplate(level, row[1].strip())

        if row[2]:
            params = row[2].strip()
            vp.set_params(params)

        if level == 1:
            valve_list.append(vp)
        elif level > vp_last.level:
            vp.parent = vp_last
            vp.add_self()
        elif level == vp_last.level:
            vp.parent = vp_last.get_parent()
            vp.add_self()
        elif level < vp_last.level:
            vp.parent = vp_last.get_parent(vp_last.level - (level - 1))
            vp.add_self()
        vp_last = vp

print(valve_list)


# загрузка доступных деталей
class Part:
    def __init__(self, tp, art_nr, name, attrs=None):
        self.tp = tp
        self.art_nr = art_nr
        self.name = name
        self.attrs = attrs

    @staticmethod
    def get_list(value):
        return list(value) if isinstance(value, Iterable) and type(value) != str else [value]

    def __repr__(self):
        return f'{self.art_nr}: {self.name}'


inventory = []

if not all(valve.name in wb.sheetnames for valve in valve_list):
    raise IndexError('Файл не содержит листа с именем соответствующим наименованию ТПА')
print(*(v.name for v in valve_list))

# дописать с учетом иерархии компонентов
parts = list(valve_list[0].bom)
print(parts)

# добавление всех деталей из листа с именем клапана в список inventory
for valve in valve_list:
    ws = wb[valve.name]
    wsl2 = tuple(tuple(cell.value for cell in row) for row in ws)
    print(wsl2)
    wsl = [row for row in wsl2 if row[0] in parts]
    print(wsl)

    for row in wsl:
        pt = Part(row[0], row[1], row[2])
        inventory.append(pt)

    for p in valve.bom:
        pass

print(*(f'{v.art_nr} {v.name}' for v in inventory), sep='\n')


wb.close()





