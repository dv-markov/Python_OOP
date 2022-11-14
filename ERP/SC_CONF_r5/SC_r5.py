# Загрузка структуры объекта и перечня компонентов из файла excel

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter
# from dataclasses import dataclass
import PySimpleGUI as sg
from datetime import datetime

VERSION = '0.5a'

wb = load_workbook("config.xlsx", read_only=True, data_only=True)
print(wb.sheetnames)
# ws = wb.active
ws = wb['Templates']
# ws = wb[wb.sheetnames[0]]

# решение через фактические значения ячеек
wsl2 = tuple(tuple(cell.value for cell in row) for row in ws)
# wb.close()
print(wsl2)
row_indx = 0
while wsl2[row_indx][0] != '#':
    row_indx += 1
print(row_indx)

# в итоговом кортеже остается только структура данных
wsl = wsl2[row_indx:]
print(wsl)


# @dataclass()
class ValvePartTemplate:
    def __init__(self, level, name, parent=None):
        self.level = level
        self.name = name
        self.parent = parent
        self.bom = dict()
        self.params = dict()

    def add_item(self, name, obj=None):
        self.bom[name] = obj

    def add_self(self):
        self.parent.add_item(self.name, self)

    def get_parent(self, depth=1):
        if depth == 1:
            return self.parent
        elif depth > 1:
            return self.parent.get_parent(depth - 1)

    def set_params(self, parameter_list):
        p = parameter_list.split(', ')
        self.params = dict.fromkeys(p)

    def get_params(self):
        return self.params

    def __repr__(self):
        return f"\n{(self.level - 1) * '-'} L:{self.level} - {self.name}" \
               f"{('; Параметры: ' + str(tuple(self.params.keys())) + '; Структура BOM: ') if self.params else ' '}" \
               f"{list(x for x in self.bom.values()) if self.bom else ''}"


# создание шаблонов BOM
valve_list = []
vp_last = None
for row in wsl:
    r = row[0].strip() if row[0] else ''
    if '#' in r:
        level = r.count('#')
        vp = ValvePartTemplate(level, row[1].strip())

        if row[2]:
            params = row[2].strip()
            vp.set_params(params)

        if level == 1:
            valve_list.append(vp)
        elif level > vp_last.level:
            vp.parent = vp_last
            vp.add_self()
        elif level == vp_last.level:
            vp.parent = vp_last.get_parent()
            vp.add_self()
        elif level < vp_last.level:
            vp.parent = vp_last.get_parent(vp_last.level - (level - 1))
            vp.add_self()
        vp_last = vp

print(valve_list)


# загрузка доступных деталей
class Part:
    def __init__(self, part_type, art_nr, name, valve_type=None, attrs=None):
        self.part_type = part_type
        self.art_nr = art_nr
        self.name = name
        self.attrs = attrs

    def __repr__(self):
        return f'{self.art_nr}: {self.name}'


inventory = []

if not all(valve.name in wb.sheetnames for valve in valve_list):
    raise IndexError('Файл не содержит листа с именем соответствующим наименованию ТПА')
print(*(v.name for v in valve_list))

# дописать с учетом иерархии компонентов
# parts = list(valve_list[0].bom)
# print(parts)

# словарь для формирования выпадающих списков
general_valve_parameters = dict()

# добавление всех деталей из листа с именем клапана в список inventory
for valve in valve_list:
    ws = wb[valve.name]
    wsl = tuple(tuple(cell.value for cell in row) for row in ws if row[0].value)

    parameter_list = {k: set() for k in valve.params}
    print(parameter_list)
    for row in wsl[1:]:
        part = Part(row[0], row[1], row[2], valve)
        part.attrs = dict(valve.bom[part.part_type].params)
        for attr in part.attrs:
            # получение списка значений параметра для данной детали
            indx = wsl[0].index(attr)
            value = row[indx]
            value_list = list(map(str.strip, value.split(';'))) if isinstance(value, str) else [str(value)]
            # запись списка значений в качестве аттрибута для данной детали
            part.attrs[attr] = value_list
            # добавление списка значений в общий словарь
            parameter_list[attr].update(value_list)

        inventory.append(part)

    # создание списка параметров для выпадающих меню (временная замена матрице)
    print(parameter_list)
    general_valve_parameters[valve.name] = {k: sorted(v) for k, v in parameter_list.items()}
    print(general_valve_parameters)

    print(f'Детали ТПА "{valve.name}" успешно загружены из файла')

print(*(f'{v.art_nr} {v.name} {v.attrs}' for v in inventory), sep='\n')
wb.close()


# начало работы с пользователем
def invite_message():
    def invite():
        print()
        print("Введите одну из указанных ниже цифр для выполнения соответствующего действия:", "\n",
              "1 - Задать конфигурацию клапана, 2 - Отобразить доступные параметры клапана, "
              "3 - Отобразить доступные детали, 0 - Выход")
        return input("> ")

    s = invite()
    while s not in ('1', '2', '3', '0'):
        print('Ошибка ввода!')
        s = invite()
    return s


def dashes(n):
    return '-' * n

def styled_cells(data):
    for c in data:
        c = Cell(ws, column="A", row=1, value=c)
        c.font = Font(bold=True)
        yield c

if __name__ == '__main__':
    title = "Конфигуратор Самсон Контролс, версия " + VERSION
    print(dashes(len(title)), title, dashes(len(title)), sep='\n')
    print()
    print(general_valve_parameters)

    sg.theme('Kayak')

    layout = [
        [sg.Text('Конфигуратор Самсон Контролс', size=(40, 2))],
        [sg.Text('')],
        [sg.Text('Выберите тип и параметры клапана')],
        [sg.Text('Тип клапана', size=(15, 1)),
         sg.Combo(list(x.name for x in valve_list), size=(20, 1), key='valve_type')]
    ]

    for key, value in general_valve_parameters['Клапан тип 3241'].items():
        layout.append([sg.Text(key, size=(15, 1)), sg.Combo(value, size=(20, 1), key=key)])

    layout.append([sg.Button('OK'), sg.Button('Выход')])

    window = sg.Window('Конфигуратор Самсон Контролс v0.5a', layout)

    wb = Workbook()
    ws = wb.active
    ws.title = valve_list[0].name
    # ws.column_dimensions['A'].width = 85
    # ws.column_dimensions['B'].width = 165
    # ws.column_dimensions['C'].width = 640
    # ws.column_dimensions['D'].width = 1500

    n = 0
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Выход':
            break
        if event == 'OK':
            print(values)
            print(values.get('valve_type'))
            valve_bom = {k: [] for k in valve_list[0].bom}
            # print(valve_list)
            # valve_bom = {k: [] for k in valve_list.get[values.get('valve_type')].bom}
            print(valve_bom)

            for part in inventory:
                if all(values[p] in part.attrs[p] for p in part.attrs):
                    valve_bom[part.part_type].append(part)

            n += 1
            # ws.append([n] + list(values.values()))
            header = list(values.values())
            header_line = [n] + [header[0]] + ['-'.join(header[1:])]
            ws.append(styled_cells(header_line))
            for key, value in valve_bom.items():
                print(key)
                ws.append([key])
                print(*(f'{v.art_nr} {v.name} {v.attrs}' for v in value), sep='\n')
                for v in value:
                    ws.append(['']+[v.art_nr, v.name, str(v.attrs)])
            ws.append([])
            sg.popup(f'BOM №{n} создан!')

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    now = datetime.now()
    dt = now.strftime("%Y%m%d-%H%M%S")
    wb.save(f"BOM {dt}.xlsx")
    window.close()










